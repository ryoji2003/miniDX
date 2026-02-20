import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

def create_excel_file(solver, shifts, staffs, tasks, days, year, month):
    """シフト表をExcel出力する（縦：職員名、横：日付、セル：業務名）"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}月シフト"

    font_header = Font(bold=True, color="FFFFFF")
    font_staff = Font(bold=True)
    fill_header = PatternFill("solid", fgColor="007BFF")
    fill_rest = PatternFill("solid", fgColor="DDDDDD")
    fill_staff_col = PatternFill("solid", fgColor="F0F4FF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # ヘッダー行: 「氏名/日付」+ 日付列
    header_cell = ws.cell(row=1, column=1, value="氏名 \\ 日付")
    header_cell.font = font_header
    header_cell.fill = fill_header
    header_cell.alignment = center
    header_cell.border = thin_border
    ws.column_dimensions['A'].width = 14

    for i, d in enumerate(days):
        col = i + 2
        date_obj = datetime.date(year, month, d)
        weekday_names = ["月", "火", "水", "木", "金", "土", "日"]
        weekday = weekday_names[date_obj.weekday()]
        c = ws.cell(row=1, column=col, value=f"{d}日\n({weekday})")
        c.font = font_header
        c.fill = fill_header
        c.alignment = center
        c.border = thin_border
        ws.column_dimensions[c.column_letter].width = 8

    # スタッフ行: スタッフ名 + 各日の業務名
    for row_idx, s in enumerate(staffs, start=2):
        # スタッフ名セル
        name_cell = ws.cell(row=row_idx, column=1, value=s.name)
        name_cell.font = font_staff
        name_cell.fill = fill_staff_col
        name_cell.alignment = center
        name_cell.border = thin_border

        for col_idx, d in enumerate(days, start=2):
            # この日に割り当てられた業務を検索
            task_name = ""
            for t in tasks:
                if solver.Value(shifts[(s.id, d, t.id)]) == 1:
                    task_name = t.name
                    break

            cell = ws.cell(row=row_idx, column=col_idx, value=task_name if task_name else "休")
            cell.alignment = center
            cell.border = thin_border

            if not task_name:
                cell.fill = fill_rest

        ws.row_dimensions[row_idx].height = 20

    ws.row_dimensions[1].height = 28

    if not os.path.exists("static"):
        os.makedirs("static")

    filename = f"static/shift_{year}_{month}_{datetime.datetime.now().strftime('%H%M%S')}.xlsx"
    wb.save(filename)
    return filename

def extract_shift_data(solver, shifts, staffs, tasks, days, year, month):
    """フロントエンド表示用にJSONデータを抽出。by_date形式とby_staff形式の両方を返す"""
    # カレンダー表示用: 日付ごとのアサインリスト
    by_date = {}
    for d in days:
        date_str = f"{year}-{month:02d}-{d:02d}"
        assignments = []
        for t in tasks:
            for s in staffs:
                if solver.Value(shifts[(s.id, d, t.id)]) == 1:
                    assignments.append({
                        "staffId": s.id,
                        "staffName": s.name,
                        "taskId": t.id,
                        "taskName": t.name,
                        "isNurse": s.is_nurse,
                    })
        if assignments:
            by_date[date_str] = assignments

    # テーブル表示用: スタッフごとの日別業務名
    by_staff = []
    for s in staffs:
        staff_shifts = {}
        for d in days:
            date_str = f"{year}-{month:02d}-{d:02d}"
            task_name = ""
            for t in tasks:
                if solver.Value(shifts[(s.id, d, t.id)]) == 1:
                    task_name = t.name
                    break
            staff_shifts[date_str] = task_name  # 空文字は休み

        by_staff.append({
            "staffId": s.id,
            "staffName": s.name,
            "shifts": staff_shifts,
        })

    return {
        "by_date": by_date,
        "by_staff": by_staff,
    }
