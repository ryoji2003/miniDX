import datetime
import os
from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

def generate_shift_excel(staffs, tasks, requirements, absences, year, month):
    """
    データベースのデータを受け取り、シフトを計算してExcelファイルのパスを返す関数
    """
    model = cp_model.CpModel()

    # --- 日付の準備 ---
    # 指定された月の日数 (28-31) を計算
    if month == 12:
        next_month = datetime.date(year + 1, 1, 1)
    else:
        next_month = datetime.date(year, month + 1, 1)

    last_day = (next_month - datetime.timedelta(days=1)).day
    days = list(range(1, last_day + 1)) # [1, 2, ..., 31]

    # --- 変数定義 ---
    # shifts[(staff_id, day, task_id)] -> 1 or 0
    shifts = {}
    for s in staffs:
        for d in days:
            for t in tasks:
                shifts[(s.id, d, t.id)] = model.NewBoolVar(f"shift_s{s.id}_d{d}_t{t.id}")

    # ==========================================
    #  制約条件 (Hard Constraints)
    # ==========================================

    # --- C1: 1日1人1業務まで (重複禁止) ---
    for s in staffs:
        for d in days:
            model.Add(sum(shifts[(s.id, d, t.id)] for t in tasks) <= 1)

    # --- C2: 日次要件 (枠) を満たす ---
    # DBの daily_requirements にある設定を守る
    # データ形式: date="YYYY-MM-DD", task_id=..., count=...
    req_map = {} # {(day, task_id): count}
    for r in requirements:
        try:
            r_date = datetime.datetime.strptime(r.date, "%Y-%m-%d")
            if r_date.year == year and r_date.month == month:
                req_map[(r_date.day, r.task_id)] = r.count
        except ValueError:
            continue # 日付形式がおかしいデータはスキップ

    for d in days:
        for t in tasks:
            if (d, t.id) in req_map:
                count = req_map[(d, t.id)]
                # その日のその業務の合計人数 == 設定人数
                model.Add(sum(shifts[(s.id, d, t.id)] for s in staffs) == count)

    # --- C3: 希望休制約 ---
    # absence_requests にある日は働かない
    abs_map = [] # [(staff_id, day)]
    for a in absences:
        try:
            a_date = datetime.datetime.strptime(a.date, "%Y-%m-%d")
            if a_date.year == year and a_date.month == month:
                abs_map.append((a.staff_id, a_date.day))
        except ValueError:
            continue

    for s_id, d in abs_map:
        # その日は全業務が0
        for t in tasks:
            # staffオブジェクトを探す（ID一致）
            # shiftsキーが存在するか確認（念のため）
            if (s_id, d, t.id) in shifts:
                 model.Add(shifts[(s_id, d, t.id)] == 0)

    # ==========================================
    #  ★W6追加ロジック (資格・免許・属性)
    # ==========================================

    # --- C4: 看護師限定業務 ---
    # 業務名に「看護」が含まれる場合、is_nurse=True の人しか入れない
    nurse_tasks = [t for t in tasks if "看護" in t.name]
    for t in nurse_tasks:
        for s in staffs:
            if not s.is_nurse:
                for d in days:
                    model.Add(shifts[(s.id, d, t.id)] == 0)

    # --- C5: 訓練限定スタッフ ---
    # can_only_train=True の人は、「訓練」業務以外は不可
    train_tasks = [t for t in tasks if "訓練" in t.name]
    if train_tasks:
        train_task_ids = [t.id for t in train_tasks]
        for s in staffs:
            if s.can_only_train:
                for d in days:
                    # 訓練以外のタスクは全て0にする
                    for t in tasks:
                        if t.id not in train_task_ids:
                             model.Add(shifts[(s.id, d, t.id)] == 0)

    # --- C6: 運転人数の確保 (送迎) ---
    # 毎日、「出勤している人」の中で、「運転できる人」が規定数以上必要
    # license_type: 1(普通), 2(ワゴン)
    # is_part_time: Trueなら運転不可

    # 普通車以上を運転できる人 (License>=1 かつ 常勤)
    all_drivers = [s for s in staffs if s.license_type >= 1 and not s.is_part_time]

    for d in days:
        # その日出勤している全ドライバー人数
        total_drivers_working = []
        for s in all_drivers:
            # そのスタッフのその日のタスク合計(0か1)
            is_working = sum(shifts[(s.id, d, t.id)] for t in tasks)
            total_drivers_working.append(is_working)

        # スタッフ不足で解なしになるのを防ぐため、人数が足りる場合のみ制約追加
        if len(all_drivers) >= 6:
            model.Add(sum(total_drivers_working) >= 6)

    # ==========================================
    #  努力目標 (Soft Constraints)
    # ==========================================

    # S1: 勤務日数の平準化 (簡易版: 上限を超えない)
    for s in staffs:
        total_work = sum(shifts[(s.id, d, t.id)] for d in days for t in tasks)
        model.Add(total_work <= s.work_limit)

    # --- ソルバー実行 ---
    solver = cp_model.CpSolver()
    # solver.parameters.max_time_in_seconds = 10.0 # 計算時間制限(必要なら)
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        excel_path = create_excel_file(solver, shifts, staffs, tasks, days, year, month)
        shift_data = extract_shift_data(solver, shifts, staffs, tasks, days, year, month)
        return excel_path, shift_data
    else:
        return None, None

def create_excel_file(solver, shifts, staffs, tasks, days, year, month):
    """計算結果からExcelを作成する"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}月シフト"

    # スタイル
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill("solid", fgColor="007BFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ヘッダー作成
    ws.cell(row=1, column=1, value="業務 / 日付")
    for i, d in enumerate(days):
        c = ws.cell(row=1, column=i+2, value=f"{d}日")
        c.font = font_header
        c.fill = fill_header
        c.alignment = center

    # データ出力
    for row_idx, t in enumerate(tasks, start=2):
        ws.cell(row=row_idx, column=1, value=t.name).alignment = center

        for col_idx, d in enumerate(days, start=2):
            assigned = []
            for s in staffs:
                if solver.Value(shifts[(s.id, d, t.id)]) == 1:
                    assigned.append(s.name)

            val = "\n".join(assigned)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center

            if not val:
                cell.fill = PatternFill("solid", fgColor="EEEEEE")

    # staticフォルダに保存
    if not os.path.exists("static"):
        os.makedirs("static")

    filename = f"static/shift_{year}_{month}_{datetime.datetime.now().strftime('%H%M%S')}.xlsx"
    wb.save(filename)
    # ファイル名だけ返す (Webからアクセスするため)
    return filename


def extract_shift_data(solver, shifts, staffs, tasks, days, year, month):
    """Extract shift data as JSON-serializable dictionary for calendar display"""
    shift_data = {}

    # Create staff lookup for quick access
    staff_map = {s.id: s for s in staffs}

    for d in days:
        date_str = f"{year}-{month:02d}-{d:02d}"
        assignments = []

        for t in tasks:
            for s in staffs:
                if solver.Value(shifts[(s.id, d, t.id)]) == 1:
                    assignments.append({
                        "staffId": s.id,
                        "staffName": s.name,
                        "taskId": t.id,
                        "taskName": t.name,
                        "isNurse": s.is_nurse,
                    })

        if assignments:
            shift_data[date_str] = assignments

    return shift_data
